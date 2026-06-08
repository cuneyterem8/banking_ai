from pathlib import Path

from openpyxl import load_workbook

from app.use_cases.aml_monitoring.data_generation import TEST_COUNT, TRAIN_COUNT, VAL_COUNT, write_artifacts
from app.use_cases.aml_monitoring.raw_data import (
    ground_truth_summary,
    load_manifest,
    load_test_alerts,
    load_train_alerts,
    load_val_alerts,
    network_summary,
    raw_artifact_paths,
)


def test_aml_data_generation_creates_expected_artifacts() -> None:
    paths = write_artifacts()

    assert Path(paths["train_xlsx"]).exists()
    assert Path(paths["val_xlsx"]).exists()
    assert Path(paths["test_xlsx"]).exists()
    assert Path(paths["transaction_network"]).exists()
    assert Path(paths["entity_relationships"]).exists()
    assert Path(paths["case_notes_pdf"]).exists()
    assert Path(paths["metadata"]).exists()
    assert Path(paths["ground_truth"]).exists()

    manifest = load_manifest()
    assert manifest["train_count"] == TRAIN_COUNT
    assert manifest["val_count"] == VAL_COUNT
    assert manifest["test_count"] == TEST_COUNT
    assert set(manifest["artifact_checksums"]) >= {
        "train_xlsx",
        "val_xlsx",
        "test_xlsx",
        "transaction_network",
        "entity_relationships",
        "case_notes_pdf",
        "ground_truth",
    }


def test_aml_raw_loaders_return_expected_counts_and_network() -> None:
    write_artifacts()

    assert len(load_train_alerts()) == TRAIN_COUNT
    assert len(load_val_alerts()) == VAL_COUNT
    assert len(load_test_alerts()) == TEST_COUNT

    summary = ground_truth_summary()
    assert summary["sar_label_counts"]["train"] > 0
    assert summary["high_risk_alert_count"] > 0

    network = network_summary()
    assert network.account_count == 850
    assert network.counterparty_count == 420
    assert network.transaction_count > 0
    assert network.entity_count == 360
    assert network.cluster_count == 36
    assert network.top_clusters


def test_aml_network_workbook_has_required_sheets() -> None:
    paths = write_artifacts()
    workbook = load_workbook(paths["transaction_network"], read_only=True, data_only=True)
    try:
      assert set(workbook.sheetnames) == {
          "accounts",
          "counterparties",
          "transactions",
          "alert_transaction_links",
      }
    finally:
      workbook.close()

    names = {path.name for path in raw_artifact_paths()}
    assert {"synthetic_aml_alerts.xlsx", "transaction_network.xlsx", "entity_relationships.json", "suspicious_activity_notes.pdf", "metadata.json", "ground_truth.json"} <= names
