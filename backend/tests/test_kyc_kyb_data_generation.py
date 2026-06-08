from pathlib import Path

import pdfplumber
from openpyxl import load_workbook

from app.use_cases.kyc_kyb.data_generation import BUSINESS_COUNT, INDIVIDUAL_COUNT, kyc_kyb_data_root, write_artifacts
from app.use_cases.kyc_kyb.raw_data import (
    ground_truth_summary,
    load_manifest,
    load_packages,
    raw_artifact_paths,
)


def test_kyc_kyb_data_generation_creates_expected_artifacts() -> None:
    paths = write_artifacts()

    assert Path(paths["metadata"]).exists()
    assert Path(paths["ground_truth"]).exists()
    assert Path(paths["sanctions_watchlist"]).exists()
    assert Path(paths["high_risk_jurisdictions"]).exists()
    assert Path(paths["document_policy_pdf"]).exists()

    manifest = load_manifest()
    assert manifest["individual_count"] == INDIVIDUAL_COUNT
    assert manifest["business_count"] == BUSINESS_COUNT
    assert manifest["package_count"] == INDIVIDUAL_COUNT + BUSINESS_COUNT
    assert manifest["document_count"] == (INDIVIDUAL_COUNT + BUSINESS_COUNT) * 6
    assert set(manifest["artifact_checksums"]) >= {
        "raw/reference/sanctions_watchlist.json",
        "raw/reference/high_risk_jurisdictions.json",
        "raw/reference/document_policy.pdf",
        "ground_truth.json",
    }


def test_kyc_kyb_raw_loaders_return_splits_and_packages() -> None:
    write_artifacts()

    train_packages = load_packages(split="train")
    val_packages = load_packages(split="val")
    test_packages = load_packages(split="test")
    assert len(train_packages) == 32
    assert len(val_packages) == 8
    assert len(test_packages) == 8
    assert {item.subject_type for item in train_packages} == {"individual", "business"}

    summary = ground_truth_summary()
    assert summary["package_count"] == 48
    assert summary["split_summary"]["train"]["package_count"] == 32
    assert summary["split_summary"]["val"]["manual_review_label_count"] > 0
    assert summary["expected_rule_flags"]


def test_kyc_kyb_generated_files_are_readable() -> None:
    write_artifacts()
    first_individual = load_packages(subject_type="individual")[0]
    proof_path = next(document for document in first_individual.documents if document.document_type == "proof_of_address")
    with pdfplumber.open(kyc_kyb_data_root() / proof_path.relative_path) as pdf:
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    assert "Synthetic Proof Of Address" in text
    assert "Customer Name:" in text

    first_business = load_packages(subject_type="business")[0]
    owners = next(document for document in first_business.documents if document.document_type == "beneficial_ownership")
    workbook = load_workbook(kyc_kyb_data_root() / owners.relative_path, read_only=True, data_only=True)
    try:
        assert workbook.sheetnames == ["beneficial_ownership"]
    finally:
        workbook.close()

    names = {path.name for path in raw_artifact_paths()}
    assert {"id_document_front.jpg", "company_registry.pdf", "document_policy.pdf", "metadata.json", "ground_truth.json"} <= names
