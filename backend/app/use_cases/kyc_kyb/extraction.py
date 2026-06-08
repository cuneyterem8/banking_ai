from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook

from app.use_cases.kyc_kyb.data_generation import kyc_kyb_data_root
from app.use_cases.kyc_kyb.llm_service import ImageFallbackClient, extract_image_fields
from app.use_cases.kyc_kyb.schemas import KycKybDocumentManifest, KycKybExtractedDocument, KycKybPackageRecord

ProgressCallback = Callable[[int, str], None]


@dataclass
class ExtractionStats:
    fallback_count: int = 0
    timeout_count: int = 0
    invalid_json_count: int = 0
    warnings: list[str] = field(default_factory=list)


def _extract_text_pdfplumber(path: Path) -> str:
    try:
        import pdfplumber
    except ModuleNotFoundError:
        return ""
    try:
        with pdfplumber.open(path) as pdf:
            return "\n".join(page.extract_text() or "" for page in pdf.pages)
    except Exception:
        return ""


def _extract_text_pymupdf(path: Path) -> str:
    try:
        import fitz
    except ModuleNotFoundError:
        return ""
    try:
        with fitz.open(path) as doc:
            return "\n".join(page.get_text("text") or "" for page in doc)
    except Exception:
        return ""


def _pdf_text(path: Path) -> str:
    text = _extract_text_pdfplumber(path)
    if len(text.strip()) >= 30:
        return text
    return _extract_text_pymupdf(path)


def _parse_label_lines(text: str) -> dict[str, str]:
    fields: dict[str, str] = {}
    for line in text.splitlines():
        if ":" not in line:
            continue
        key, value = line.split(":", 1)
        if key.strip() and value.strip():
            fields[key.strip()] = value.strip()
    return fields


def _xlsx_fields(path: Path, document_type: str) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if len(rows) < 2:
        return {}
    headers = [str(item) for item in rows[0]]
    records = [
        {header: value for header, value in zip(headers, row)}
        for row in rows[1:]
        if any(value is not None for value in row)
    ]
    if document_type == "beneficial_ownership":
        valid_owner_records = [
            record
            for record in records
            if str(record.get("Owner Name") or "").strip()
            and float(record.get("Ownership Percent") or 0) > 0
        ]
        return {
            "Owner Count": len(valid_owner_records),
            "Ownership Total": sum(float(record.get("Ownership Percent") or 0) for record in valid_owner_records),
            "Owners": valid_owner_records,
        }
    return dict(records[0]) if records else {}


def _normalize(value: Any) -> str:
    return str(value).strip().lower()


def _validate_fields(expected: dict[str, Any], actual: dict[str, Any]) -> tuple[float, list[str]]:
    if not expected:
        return 1.0, []
    matched = 0
    issues: list[str] = []
    for key, expected_value in expected.items():
        actual_value = actual.get(key)
        if actual_value is None:
            issues.append(f"Missing field: {key}")
            continue
        if isinstance(expected_value, (int, float)):
            try:
                if abs(float(actual_value) - float(expected_value)) < 0.01:
                    matched += 1
                else:
                    issues.append(f"Mismatch for {key}: expected {expected_value}, got {actual_value}")
            except (TypeError, ValueError):
                issues.append(f"Mismatch for {key}: expected {expected_value}, got {actual_value}")
            continue
        if _normalize(actual_value) == _normalize(expected_value):
            matched += 1
        else:
            issues.append(f"Mismatch for {key}: expected {expected_value}, got {actual_value}")
    return matched / len(expected), issues


def _extract_document(
    document: KycKybDocumentManifest,
    *,
    fallback_client: ImageFallbackClient | None = None,
) -> tuple[KycKybExtractedDocument, ExtractionStats]:
    path = kyc_kyb_data_root() / document.relative_path
    stats = ExtractionStats()
    provider = "local-ocr"
    status = "completed"
    raw_text = ""
    fields: dict[str, Any] = {}
    if document.is_image:
        result = extract_image_fields(document, path, fallback_client=fallback_client)
        fields = result.fields
        provider = result.provider_used
        status = result.status
        raw_text = result.raw_text_excerpt
        stats.fallback_count += result.fallback_count
        stats.timeout_count += result.timeout_count
        stats.invalid_json_count += result.invalid_json_count
        stats.warnings.extend(result.warnings)
    elif path.suffix.lower() == ".pdf":
        raw_text = _pdf_text(path)
        fields = _parse_label_lines(raw_text)
        provider = "local-ocr"
    elif path.suffix.lower() == ".xlsx":
        fields = _xlsx_fields(path, document.document_type)
        raw_text = " ".join(f"{key}: {value}" for key, value in fields.items() if key != "Owners")
        provider = "local-openpyxl"
    field_score, issues = _validate_fields(document.expected_fields, fields)
    if status == "fallback_unavailable":
        issues.append("Image extraction fallback unavailable.")
    confidence = 0.0 if status == "fallback_unavailable" else round(field_score, 4)
    extracted = KycKybExtractedDocument(
        package_id=document.package_id,
        document_id=document.document_id,
        document_type=document.document_type,
        file_name=document.file_name,
        provider_used=provider,
        extraction_status=status,
        confidence=confidence,
        fields=fields,
        validation_issues=issues[:10],
        raw_text_excerpt=" ".join(str(raw_text).split())[:500],
    )
    return extracted, stats


def extract_packages(
    packages: list[KycKybPackageRecord],
    *,
    fallback_client: ImageFallbackClient | None = None,
    progress_callback: ProgressCallback | None = None,
) -> tuple[list[KycKybExtractedDocument], ExtractionStats]:
    documents = [document for package in packages for document in package.documents]
    extracted: list[KycKybExtractedDocument] = []
    aggregate = ExtractionStats()
    for index, document in enumerate(documents, start=1):
        if progress_callback:
            progress_callback(5 + int((index - 1) / max(len(documents), 1) * 35), f"extracting_{document.document_type}")
        item, stats = _extract_document(document, fallback_client=fallback_client)
        extracted.append(item)
        aggregate.fallback_count += stats.fallback_count
        aggregate.timeout_count += stats.timeout_count
        aggregate.invalid_json_count += stats.invalid_json_count
        aggregate.warnings.extend(stats.warnings)
    aggregate.warnings = list(dict.fromkeys(aggregate.warnings))
    return extracted, aggregate
