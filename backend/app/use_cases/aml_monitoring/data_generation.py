from __future__ import annotations

import hashlib
import json
import math
import random
import shutil
import textwrap
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

from app.data_paths import get_use_case_data_dir
from app.use_cases.aml_monitoring.feature_engineering import ALERT_TYPE_RISK, TYPOLOGY_RISK

USE_CASE_SLUG = "aml-monitoring"
FILE_BASENAME = "synthetic_aml_alerts"
GENERATION_SEED = 9306
TRAIN_COUNT = 1600
VAL_COUNT = 400
TEST_COUNT = 500
TOTAL_COUNT = TRAIN_COUNT + VAL_COUNT + TEST_COUNT
LABEL_NOISE_RATE = 0.01

HIGH_RISK_JURISDICTIONS = ["Northland Free Zone", "Orchid Islands", "Silver Coast", "Eastport Trade Hub"]
REGIONS = ["Metro", "North", "South", "West", "East", "Central"]
ALERT_TYPES = list(ALERT_TYPE_RISK)
TYPOLOGIES = list(TYPOLOGY_RISK)
RULE_TRIGGERS = [
    "cash deposits below reporting threshold",
    "rapid outgoing wire after cash deposit",
    "counterparty in high-risk jurisdiction",
    "beneficial owner mismatch",
    "sanctions name similarity",
    "adverse media keyword match",
    "round amount transaction pattern",
    "nested entity relationship",
    "unusual overnight transaction time",
]
HEADERS = [
    "alert_id",
    "customer_id",
    "account_id",
    "entity_id",
    "kyc_risk_score",
    "jurisdiction_risk_score",
    "sanctions_name_similarity",
    "adverse_media_flag",
    "prior_alert_count_12m",
    "cash_deposit_total_30d",
    "outgoing_wire_total_30d",
    "round_amount_ratio",
    "rapid_movement_ratio",
    "structuring_count_7d",
    "unusual_hours_count",
    "counterparty_cluster_risk",
    "network_degree",
    "network_centrality_score",
    "nested_entity_depth",
    "beneficial_owner_mismatch",
    "alert_type",
    "typology_tag",
    "rule_triggers",
    "linked_transaction_count",
    "related_entities",
    "label_sar_recommended",
]


def aml_data_root() -> Path:
    return get_use_case_data_dir(USE_CASE_SLUG)


def aml_raw_root() -> Path:
    return aml_data_root() / "raw"


def aml_split_dir(split: str) -> Path:
    return aml_raw_root() / split


def aml_xlsx_path(split: str) -> Path:
    return aml_split_dir(split) / f"{FILE_BASENAME}.xlsx"


def transaction_network_path() -> Path:
    return aml_raw_root() / "network" / "transaction_network.xlsx"


def entity_relationships_path() -> Path:
    return aml_raw_root() / "entities" / "entity_relationships.json"


def case_notes_pdf_path() -> Path:
    return aml_raw_root() / "cases" / "suspicious_activity_notes.pdf"


def metadata_path() -> Path:
    return aml_data_root() / "metadata.json"


def ground_truth_path() -> Path:
    return aml_data_root() / "ground_truth.json"


def _sigmoid(value: float) -> float:
    if value >= 0:
        z = math.exp(-value)
        return 1 / (1 + z)
    z = math.exp(value)
    return z / (1 + z)


def _clamp(value: float, low: float, high: float) -> float:
    return max(low, min(high, value))


def _checksum(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _sample_alert(index: int, rng: random.Random) -> dict[str, Any]:
    alert_type = rng.choices(ALERT_TYPES, weights=[22, 18, 12, 12, 16, 10], k=1)[0]
    typology = rng.choices(TYPOLOGIES, weights=[20, 16, 10, 8, 10, 14, 22], k=1)[0]
    base_risk = 0.36 * TYPOLOGY_RISK[typology] + 0.25 * ALERT_TYPE_RISK[alert_type] + rng.uniform(-0.08, 0.1)
    kyc = _clamp(rng.betavariate(2.3, 3.2) * 0.65 + base_risk * 0.35, 0.02, 0.98)
    jurisdiction = _clamp(rng.betavariate(1.7, 3.5) * 0.7 + (0.28 if typology == "High-Risk Jurisdiction" else 0), 0.02, 0.98)
    sanctions = _clamp(rng.betavariate(1.2, 8.5) + (0.48 if typology == "Sanctions Proximity" else 0), 0.0, 0.99)
    adverse = 1 if rng.random() < (0.14 + (0.48 if typology == "Adverse Media" else 0.0)) else 0
    structuring = rng.choices(range(0, 10), weights=[18, 16, 14, 12, 10, 8, 7, 6, 4, 3], k=1)[0]
    if typology == "Structuring":
        structuring += rng.randint(2, 5)
    rapid = _clamp(rng.betavariate(1.8, 3.8) + (0.28 if typology == "Rapid Movement" else 0), 0.02, 0.99)
    cluster = _clamp(rng.betavariate(1.9, 3.4) + (0.3 if typology == "Shell Entity Network" else 0), 0.02, 0.99)
    network_degree = rng.randint(2, 18) + (rng.randint(5, 14) if typology == "Shell Entity Network" else 0)
    nested_depth = rng.choices([0, 1, 2, 3, 4, 5], weights=[24, 24, 20, 16, 10, 6], k=1)[0]
    if typology == "Shell Entity Network":
        nested_depth = max(nested_depth, rng.randint(3, 5))
    owner_mismatch = 1 if rng.random() < (0.08 + 0.32 * cluster + (0.25 if nested_depth >= 3 else 0)) else 0
    cash_total = round(rng.lognormvariate(10.1, 0.75) * (1.35 if typology in {"Structuring", "Unusual Cash Activity"} else 0.8), 2)
    wire_total = round(rng.lognormvariate(10.0, 0.8) * (1.45 if typology in {"Rapid Movement", "High-Risk Jurisdiction"} else 0.75), 2)
    round_ratio = _clamp(rng.betavariate(1.4, 4.5) + (0.32 if typology == "Structuring" else 0), 0.01, 0.98)
    unusual_hours = rng.randint(0, 7) + (rng.randint(1, 5) if rapid > 0.65 else 0)
    linked_transactions = rng.randint(3, 26) + (structuring * 2) + (network_degree // 3)
    prior_alerts = rng.choices(range(0, 8), weights=[32, 22, 16, 11, 8, 5, 3, 2], k=1)[0]
    trigger_pool = list(RULE_TRIGGERS)
    rng.shuffle(trigger_pool)
    trigger_count = max(1, min(len(trigger_pool), 1 + int(structuring >= 4) + int(rapid >= 0.55) + int(cluster >= 0.6) + int(sanctions >= 0.45) + int(adverse)))
    related_count = max(1, min(5, network_degree // 6 + 1))
    related_entities = [f"ENT-{2000 + ((index + offset * 17) % 360):04d}" for offset in range(related_count)]
    row = {
        "alert_id": f"AML-{index + 1:06d}",
        "customer_id": f"CUST-{7000 + (index % 520):04d}",
        "account_id": f"ACCT-{90000 + (index % 850):05d}",
        "entity_id": f"ENT-{2000 + (index % 360):04d}",
        "kyc_risk_score": round(kyc, 4),
        "jurisdiction_risk_score": round(jurisdiction, 4),
        "sanctions_name_similarity": round(sanctions, 4),
        "adverse_media_flag": adverse,
        "prior_alert_count_12m": prior_alerts,
        "cash_deposit_total_30d": cash_total,
        "outgoing_wire_total_30d": wire_total,
        "round_amount_ratio": round(round_ratio, 4),
        "rapid_movement_ratio": round(rapid, 4),
        "structuring_count_7d": structuring,
        "unusual_hours_count": unusual_hours,
        "counterparty_cluster_risk": round(cluster, 4),
        "network_degree": network_degree,
        "network_centrality_score": round(_clamp(cluster * rng.uniform(0.55, 1.08), 0.01, 0.99), 4),
        "nested_entity_depth": nested_depth,
        "beneficial_owner_mismatch": owner_mismatch,
        "alert_type": alert_type,
        "typology_tag": typology,
        "rule_triggers": "; ".join(trigger_pool[:trigger_count]),
        "linked_transaction_count": linked_transactions,
        "related_entities": ";".join(related_entities),
    }
    logit = (
        -3.05
        + 1.05 * row["kyc_risk_score"]
        + 1.1 * row["jurisdiction_risk_score"]
        + 1.7 * row["sanctions_name_similarity"]
        + 0.72 * row["adverse_media_flag"]
        + 0.12 * row["prior_alert_count_12m"]
        + 0.16 * min(row["structuring_count_7d"], 9)
        + 1.1 * row["rapid_movement_ratio"]
        + 1.05 * row["counterparty_cluster_risk"]
        + 0.82 * row["beneficial_owner_mismatch"]
        + 0.16 * row["nested_entity_depth"]
        + rng.gauss(0, 0.28)
    )
    label = 1 if rng.random() < _sigmoid(logit) else 0
    if rng.random() < LABEL_NOISE_RATE:
        label = 1 - label
    row["label_sar_recommended"] = label
    return row


def build_alerts(count: int = TOTAL_COUNT, seed: int = GENERATION_SEED) -> list[dict[str, Any]]:
    rng = random.Random(seed)
    return [_sample_alert(index, rng) for index in range(count)]


def split_alerts(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    rng = random.Random(GENERATION_SEED + 41)
    positive = [row for row in rows if row["label_sar_recommended"] == 1]
    negative = [row for row in rows if row["label_sar_recommended"] == 0]
    rng.shuffle(positive)
    rng.shuffle(negative)
    rate = len(positive) / len(rows)

    def take(count: int) -> list[dict[str, Any]]:
        nonlocal positive, negative
        positive_count = min(len(positive), max(1, round(count * rate)))
        negative_count = count - positive_count
        chunk = positive[:positive_count] + negative[:negative_count]
        positive = positive[positive_count:]
        negative = negative[negative_count:]
        rng.shuffle(chunk)
        return chunk

    test = take(TEST_COUNT)
    val = take(VAL_COUNT)
    train = take(TRAIN_COUNT)
    return train, val, test


def _reset_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)
    for item in path.iterdir():
        if item.name == ".gitkeep":
            continue
        if item.is_dir():
            shutil.rmtree(item)
        else:
            item.unlink()


def _write_split_xlsx(split: str, rows: list[dict[str, Any]]) -> str:
    split_dir = aml_split_dir(split)
    _reset_dir(split_dir)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"AML_{split}"[:31]
    sheet.append(HEADERS)
    for row in rows:
        sheet.append([row[header] for header in HEADERS])
    for column in sheet.columns:
        letter = column[0].column_letter
        sheet.column_dimensions[letter].width = min(max(len(str(cell.value or "")) for cell in column) + 2, 34)
    workbook.save(aml_xlsx_path(split))
    return str(aml_xlsx_path(split).resolve())


def _write_network_workbook(rows: list[dict[str, Any]]) -> str:
    path = transaction_network_path()
    _reset_dir(path.parent)
    workbook = Workbook()
    accounts = workbook.active
    accounts.title = "accounts"
    accounts.append(["account_id", "customer_id", "entity_id", "region", "account_type", "open_months", "base_risk_score"])
    for index, row in enumerate(rows[:850]):
        accounts.append([
            row["account_id"],
            row["customer_id"],
            row["entity_id"],
            REGIONS[index % len(REGIONS)],
            "business" if index % 5 == 0 else "retail",
            3 + (index % 96),
            row["kyc_risk_score"],
        ])
    counterparties = workbook.create_sheet("counterparties")
    counterparties.append(["counterparty_id", "counterparty_name", "jurisdiction", "cluster_id", "cluster_risk_score"])
    for index in range(420):
        jurisdiction = HIGH_RISK_JURISDICTIONS[index % len(HIGH_RISK_JURISDICTIONS)] if index % 9 == 0 else f"Synthetic Country {index % 17}"
        counterparties.append([
            f"CP-{3000 + index:04d}",
            f"Synthetic Counterparty {index + 1}",
            jurisdiction,
            f"CL-{1 + index % 36:03d}",
            round(0.18 + (index % 11) * 0.055 + (0.22 if index % 9 == 0 else 0), 4),
        ])
    transactions = workbook.create_sheet("transactions")
    transactions.append(["transaction_id", "account_id", "counterparty_id", "amount", "direction", "channel", "jurisdiction", "days_before_alert"])
    for index, row in enumerate(rows[:2200]):
        for offset in range(2):
            cp_index = (index * 3 + offset * 19) % 420
            high_risk = cp_index % 9 == 0
            transactions.append([
                f"AML-TXN-{index + 1:06d}-{offset + 1}",
                row["account_id"],
                f"CP-{3000 + cp_index:04d}",
                round((row["cash_deposit_total_30d"] + row["outgoing_wire_total_30d"]) / max(row["linked_transaction_count"], 1) * (0.7 + offset * 0.35), 2),
                "outgoing" if offset else "incoming",
                "wire" if row["outgoing_wire_total_30d"] > row["cash_deposit_total_30d"] else "cash",
                HIGH_RISK_JURISDICTIONS[cp_index % len(HIGH_RISK_JURISDICTIONS)] if high_risk else f"Synthetic Country {cp_index % 17}",
                (index + offset) % 30,
            ])
    links = workbook.create_sheet("alert_transaction_links")
    links.append(["alert_id", "transaction_id", "link_reason"])
    for index, row in enumerate(rows):
        links.append([row["alert_id"], f"AML-TXN-{index % 2200 + 1:06d}-1", "primary linked transaction"])
        if row["linked_transaction_count"] > 10:
            links.append([row["alert_id"], f"AML-TXN-{index % 2200 + 1:06d}-2", "secondary linked transaction"])
    workbook.save(path)
    return str(path.resolve())


def _write_entity_relationships(rows: list[dict[str, Any]]) -> str:
    path = entity_relationships_path()
    _reset_dir(path.parent)
    entities = []
    clusters = []
    for index in range(360):
        high_risk = index % 11 == 0
        entity_id = f"ENT-{2000 + index:04d}"
        entities.append(
            {
                "entity_id": entity_id,
                "legal_name": f"Synthetic Entity {index + 1} LLC",
                "beneficial_owner_id": f"BO-{500 + index % 140:04d}",
                "beneficial_owner_name": f"Synthetic Owner {index % 140 + 1}",
                "jurisdiction": HIGH_RISK_JURISDICTIONS[index % len(HIGH_RISK_JURISDICTIONS)] if high_risk else f"Synthetic Country {index % 17}",
                "nested_entity_depth": index % 6,
                "related_accounts": [row["account_id"] for row in rows if row["entity_id"] == entity_id][:6],
                "high_risk": high_risk,
            }
        )
    for index in range(36):
        clusters.append(
            {
                "cluster_id": f"CL-{index + 1:03d}",
                "cluster_name": f"Synthetic AML Cluster {index + 1}",
                "entity_ids": [f"ENT-{2000 + ((index * 10 + offset) % 360):04d}" for offset in range(6)],
                "cluster_risk_score": round(0.22 + (index % 9) * 0.075 + (0.18 if index % 5 == 0 else 0), 4),
                "dominant_typology": TYPOLOGIES[index % len(TYPOLOGIES)],
            }
        )
    payload = {
        "generation_seed": GENERATION_SEED,
        "high_risk_jurisdictions": HIGH_RISK_JURISDICTIONS,
        "entities": entities,
        "clusters": clusters,
    }
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return str(path.resolve())


def _write_case_notes_pdf() -> str:
    path = case_notes_pdf_path()
    _reset_dir(path.parent)
    doc = canvas.Canvas(str(path), pagesize=letter)
    width, height = letter
    y = height - 54
    doc.setFont("Helvetica-Bold", 14)
    doc.drawString(54, y, "Synthetic AML Suspicious Activity Notes")
    y -= 30
    doc.setFont("Helvetica", 9)
    paragraphs = [
        "These notes are generated for a local MVP. They contain no real customer or bank data.",
        "Escalation example: repeated cash deposits below reporting thresholds followed by rapid outgoing wires should be reviewed for structuring and layering indicators.",
        "Escalation example: sanctions name similarity above the internal review band requires analyst review even when the transaction amount is moderate.",
        "Escalation example: nested entities with beneficial owner mismatch should be connected to entity relationship evidence before drafting a SAR narrative.",
        "Narrative guidance: summarize the observed pattern, name the synthetic typology, cite concrete signals, and list missing information without making legal conclusions.",
    ]
    note_id = 1
    for paragraph in paragraphs:
        for line in textwrap.wrap(paragraph, width=96):
            doc.drawString(54, y, line)
            y -= 14
        y -= 10
    for typology in TYPOLOGIES:
        text = f"Case note {note_id}: {typology} alerts should include transaction pattern, customer risk context, related entities, and recommended next steps."
        for line in textwrap.wrap(text, width=96):
            doc.drawString(54, y, line)
            y -= 14
        y -= 6
        note_id += 1
        if y < 72:
            doc.showPage()
            doc.setFont("Helvetica", 9)
            y = height - 54
    doc.save()
    return str(path.resolve())


def _ground_truth(train_rows: list[dict[str, Any]], val_rows: list[dict[str, Any]], test_rows: list[dict[str, Any]]) -> dict[str, Any]:
    all_rows = train_rows + val_rows + test_rows
    high_risk = [
        row["alert_id"]
        for row in all_rows
        if row["label_sar_recommended"] == 1 and (row["sanctions_name_similarity"] >= 0.6 or row["counterparty_cluster_risk"] >= 0.7)
    ][:60]
    return {
        "generation_seed": GENERATION_SEED,
        "train_count": len(train_rows),
        "val_count": len(val_rows),
        "test_count": len(test_rows),
        "sar_label_counts": {
            "train": sum(row["label_sar_recommended"] for row in train_rows),
            "val": sum(row["label_sar_recommended"] for row in val_rows),
            "test": sum(row["label_sar_recommended"] for row in test_rows),
        },
        "expected_typologies": TYPOLOGIES,
        "high_risk_alert_ids": high_risk,
        "narrative_evidence_snippets": [
            "cash deposits below reporting threshold",
            "rapid outgoing wire after cash deposit",
            "beneficial owner mismatch",
            "sanctions name similarity",
            "counterparty in high-risk jurisdiction",
        ],
    }


def write_artifacts() -> dict[str, str]:
    root = aml_data_root()
    root.mkdir(parents=True, exist_ok=True)
    rows = build_alerts()
    train_rows, val_rows, test_rows = split_alerts(rows)
    paths = {
        "train_xlsx": _write_split_xlsx("train", train_rows),
        "val_xlsx": _write_split_xlsx("val", val_rows),
        "test_xlsx": _write_split_xlsx("test", test_rows),
        "transaction_network": _write_network_workbook(rows),
        "entity_relationships": _write_entity_relationships(rows),
        "case_notes_pdf": _write_case_notes_pdf(),
    }
    ground_truth = _ground_truth(train_rows, val_rows, test_rows)
    ground_truth_path().write_text(json.dumps(ground_truth, indent=2), encoding="utf-8")
    paths["ground_truth"] = str(ground_truth_path().resolve())
    artifacts = {key: _checksum(Path(value)) for key, value in paths.items()}
    metadata = {
        "dataset": FILE_BASENAME,
        "generation_seed": GENERATION_SEED,
        "total_generated_rows": len(rows),
        "train_count": len(train_rows),
        "val_count": len(val_rows),
        "test_count": len(test_rows),
        "train_sar_label_count": ground_truth["sar_label_counts"]["train"],
        "val_sar_label_count": ground_truth["sar_label_counts"]["val"],
        "test_sar_label_count": ground_truth["sar_label_counts"]["test"],
        "label_noise_rate": LABEL_NOISE_RATE,
        "columns": [column for column in HEADERS if column != "label_sar_recommended"],
        "expected_typologies": TYPOLOGIES,
        "high_risk_jurisdictions": HIGH_RISK_JURISDICTIONS,
        "artifact_checksums": artifacts,
        "description": "Synthetic AML alerts with risk, transaction behavior, network, entity, and typology fields for SAR prioritization.",
    }
    metadata_path().write_text(json.dumps(metadata, indent=2), encoding="utf-8")
    paths["metadata"] = str(metadata_path().resolve())
    return paths
