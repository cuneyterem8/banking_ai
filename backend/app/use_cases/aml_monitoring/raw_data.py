from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from app.use_cases.aml_monitoring.data_generation import (
    TEST_COUNT,
    TRAIN_COUNT,
    VAL_COUNT,
    aml_data_root,
    aml_raw_root,
    aml_xlsx_path,
    case_notes_pdf_path,
    entity_relationships_path,
    ground_truth_path,
    metadata_path,
    transaction_network_path,
    write_artifacts,
)
from app.use_cases.aml_monitoring.schemas import AmlAlertRecord, AmlCaseNoteSummary, AmlNetworkSummary

USE_CASE_SLUG = "aml-monitoring"
DATASET_KEY_TRAIN = "train"
DATASET_KEY_VAL = "val"
DATASET_KEY_TEST = "test"
DATASET_KEY_NETWORK = "network"
DATASET_KEY_ENTITIES = "entities"
DATASET_KEY_CASE_NOTES = "case_notes"


def ensure_raw_artifacts() -> None:
    required = [
        aml_xlsx_path("train"),
        aml_xlsx_path("val"),
        aml_xlsx_path("test"),
        transaction_network_path(),
        entity_relationships_path(),
        case_notes_pdf_path(),
        metadata_path(),
        ground_truth_path(),
    ]
    if not all(path.exists() for path in required):
        write_artifacts()
        return
    manifest = load_manifest(regenerate_if_missing=False)
    if (
        manifest.get("train_count") != TRAIN_COUNT
        or manifest.get("val_count") != VAL_COUNT
        or manifest.get("test_count") != TEST_COUNT
    ):
        write_artifacts()


def load_manifest(*, regenerate_if_missing: bool = True) -> dict[str, Any]:
    if regenerate_if_missing and not metadata_path().exists():
        write_artifacts()
    return json.loads(metadata_path().read_text(encoding="utf-8"))


def load_ground_truth() -> dict[str, Any]:
    ensure_raw_artifacts()
    return json.loads(ground_truth_path().read_text(encoding="utf-8"))


def _load_from_xlsx(split: str) -> list[AmlAlertRecord]:
    ensure_raw_artifacts()
    frame = pd.read_excel(aml_xlsx_path(split))
    return [AmlAlertRecord(**record) for record in frame.to_dict(orient="records")]


def load_train_alerts() -> list[AmlAlertRecord]:
    return _load_from_xlsx("train")


def load_val_alerts() -> list[AmlAlertRecord]:
    return _load_from_xlsx("val")


def load_test_alerts() -> list[AmlAlertRecord]:
    return _load_from_xlsx("test")


def load_entity_relationships() -> dict[str, Any]:
    ensure_raw_artifacts()
    return json.loads(entity_relationships_path().read_text(encoding="utf-8"))


def raw_artifact_paths() -> list[Path]:
    ensure_raw_artifacts()
    artifact_paths = sorted(path for path in aml_raw_root().rglob("*") if path.is_file())
    return artifact_paths + [metadata_path(), ground_truth_path()]


def manifest_preview(limit: int = 12) -> list[dict[str, Any]]:
    return [item.model_dump() for item in load_val_alerts()[:limit]]


def ground_truth_summary() -> dict[str, Any]:
    ground_truth = load_ground_truth()
    return {
        "train_count": ground_truth["train_count"],
        "val_count": ground_truth["val_count"],
        "test_count": ground_truth["test_count"],
        "sar_label_counts": ground_truth["sar_label_counts"],
        "expected_typologies": ground_truth["expected_typologies"],
        "high_risk_alert_count": len(ground_truth["high_risk_alert_ids"]),
        "narrative_evidence_snippets": ground_truth["narrative_evidence_snippets"],
    }


def network_summary() -> AmlNetworkSummary:
    ensure_raw_artifacts()
    workbook = load_workbook(transaction_network_path(), read_only=True, data_only=True)
    try:
        account_count = max(workbook["accounts"].max_row - 1, 0)
        counterparty_count = max(workbook["counterparties"].max_row - 1, 0)
        transaction_count = max(workbook["transactions"].max_row - 1, 0)
        alert_link_count = max(workbook["alert_transaction_links"].max_row - 1, 0)
    finally:
        workbook.close()

    relationships = load_entity_relationships()
    clusters = relationships.get("clusters", [])
    high_risk_clusters = sorted(
        clusters,
        key=lambda item: float(item.get("cluster_risk_score", 0)),
        reverse=True,
    )[:5]
    return AmlNetworkSummary(
        account_count=account_count,
        counterparty_count=counterparty_count,
        transaction_count=transaction_count,
        alert_link_count=alert_link_count,
        entity_count=len(relationships.get("entities", [])),
        cluster_count=len(clusters),
        high_risk_cluster_count=sum(1 for item in clusters if float(item.get("cluster_risk_score", 0)) >= 0.7),
        high_risk_jurisdictions=[str(item) for item in relationships.get("high_risk_jurisdictions", [])],
        top_clusters=[
            {
                "cluster_id": item.get("cluster_id"),
                "cluster_name": item.get("cluster_name"),
                "cluster_risk_score": item.get("cluster_risk_score"),
                "dominant_typology": item.get("dominant_typology"),
                "entity_count": len(item.get("entity_ids", [])),
            }
            for item in high_risk_clusters
        ],
    )


def case_note_summary() -> AmlCaseNoteSummary:
    ensure_raw_artifacts()
    text = "Synthetic AML Suspicious Activity Notes"
    try:
        import pypdf

        reader = pypdf.PdfReader(str(case_notes_pdf_path()))
        text = "\n".join(page.extract_text() or "" for page in reader.pages)
    except Exception:
        pass
    keywords = [
        "structuring",
        "rapid outgoing wire",
        "beneficial owner mismatch",
        "sanctions name similarity",
        "nested entities",
    ]
    return AmlCaseNoteSummary(
        file_name=case_notes_pdf_path().name,
        note_count=max(1, text.count("Case note")),
        escalation_topic_count=sum(1 for keyword in keywords if keyword.lower() in text.lower()),
        guidance_excerpt=text[:520],
    )


def aml_data_relative(path: Path) -> str:
    return str(path.resolve().relative_to(aml_data_root().resolve())).replace("\\", "/")
