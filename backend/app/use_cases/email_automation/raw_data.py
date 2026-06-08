from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.use_cases.email_automation.data_generation import (
    CAMPAIGN_AUDIENCE_COUNT,
    CUSTOMER_COUNT,
    EVALUATION_CASE_COUNT,
    SERVICE_EVENT_COUNT,
    campaign_plan_path,
    campaign_templates_path,
    customer_events_path,
    customer_profiles_path,
    email_data_root,
    email_raw_root,
    evaluation_cases_path,
    ground_truth_path,
    metadata_path,
    service_templates_path,
    write_artifacts,
)
from app.use_cases.email_automation.schemas import (
    CampaignRecord,
    CustomerEvent,
    CustomerProfile,
    EmailGenerationCase,
    EmailTemplate,
)

USE_CASE_SLUG = "email-automation"
DATASET_KEY_EMAIL_INPUTS = "email_generation_inputs"


def ensure_raw_artifacts() -> None:
    required = [
        metadata_path(),
        ground_truth_path(),
        customer_profiles_path(),
        customer_events_path(),
        campaign_plan_path(),
        service_templates_path(),
        campaign_templates_path(),
        evaluation_cases_path(),
    ]
    if not all(path.exists() for path in required):
        write_artifacts()
        return
    manifest = load_manifest(regenerate_if_missing=False)
    if (
        manifest.get("customer_count") != CUSTOMER_COUNT
        or manifest.get("service_event_count") != SERVICE_EVENT_COUNT
        or manifest.get("campaign_audience_count") != CAMPAIGN_AUDIENCE_COUNT
        or manifest.get("evaluation_case_count") != EVALUATION_CASE_COUNT
    ):
        write_artifacts()


def load_manifest(*, regenerate_if_missing: bool = True) -> dict[str, Any]:
    if regenerate_if_missing and not metadata_path().exists():
        write_artifacts()
    return json.loads(metadata_path().read_text(encoding="utf-8"))


def load_ground_truth(*, regenerate_if_missing: bool = True) -> dict[str, Any]:
    if regenerate_if_missing:
        ensure_raw_artifacts()
    return json.loads(ground_truth_path().read_text(encoding="utf-8"))


def _load_xlsx_records(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if len(rows) < 2:
        return []
    headers = [str(item) for item in rows[0]]
    return [
        {header: value for header, value in zip(headers, row)}
        for row in rows[1:]
        if any(value is not None for value in row)
    ]


def load_customers() -> list[CustomerProfile]:
    ensure_raw_artifacts()
    return [CustomerProfile(**item) for item in _load_xlsx_records(customer_profiles_path())]


def load_events() -> list[CustomerEvent]:
    ensure_raw_artifacts()
    return [CustomerEvent(**item) for item in json.loads(customer_events_path().read_text(encoding="utf-8"))]


def load_campaigns() -> list[CampaignRecord]:
    ensure_raw_artifacts()
    return [CampaignRecord(**item) for item in _load_xlsx_records(campaign_plan_path())]


def _parse_template_file(path: Path) -> list[EmailTemplate]:
    templates: list[EmailTemplate] = []
    current: dict[str, Any] | None = None
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line:
            if current:
                current["required_disclosures"] = [
                    item.strip()
                    for item in str(current.get("required_disclosures") or "").split("|")
                    if item.strip()
                ]
                templates.append(EmailTemplate(**current))
                current = None
            continue
        if line.startswith("[") and line.endswith("]"):
            if current:
                current["required_disclosures"] = [
                    item.strip()
                    for item in str(current.get("required_disclosures") or "").split("|")
                    if item.strip()
                ]
                templates.append(EmailTemplate(**current))
            current = {"template_key": line[1:-1]}
            continue
        if current is not None and "=" in line:
            key, value = line.split("=", 1)
            current[key.strip()] = value.strip()
    if current:
        current["required_disclosures"] = [
            item.strip()
            for item in str(current.get("required_disclosures") or "").split("|")
            if item.strip()
        ]
        templates.append(EmailTemplate(**current))
    return templates


def load_templates() -> list[EmailTemplate]:
    ensure_raw_artifacts()
    return [*_parse_template_file(service_templates_path()), *_parse_template_file(campaign_templates_path())]


def load_evaluation_cases() -> list[EmailGenerationCase]:
    ensure_raw_artifacts()
    return [EmailGenerationCase(**item) for item in json.loads(evaluation_cases_path().read_text(encoding="utf-8"))]


def manifest_preview(limit: int = 16) -> list[dict[str, Any]]:
    customers = load_customers()
    events = load_events()
    campaigns = load_campaigns()
    cases = load_evaluation_cases()
    preview = [
        {"record_type": "customer", "id": item.customer_id, "segment": item.segment, "marketing_opt_in": item.marketing_opt_in}
        for item in customers[:4]
    ] + [
        {"record_type": "event", "id": item.event_id, "customer_id": item.customer_id, "event_type": item.event_type}
        for item in events[:4]
    ] + [
        {"record_type": "campaign", "id": item.audience_id, "customer_id": item.customer_id, "campaign_type": item.campaign_type}
        for item in campaigns[:4]
    ] + [
        {"record_type": "case", "id": item.case_id, "customer_id": item.customer_id, "communication_type": item.communication_type}
        for item in cases[:4]
    ]
    return preview[:limit]


def ground_truth_summary() -> dict[str, Any]:
    ground_truth = load_ground_truth()
    return {
        "customer_count": ground_truth["customer_count"],
        "service_event_count": ground_truth["service_event_count"],
        "campaign_audience_count": ground_truth["campaign_audience_count"],
        "evaluation_case_count": ground_truth["evaluation_case_count"],
        "required_rule_ids": ground_truth["required_rule_ids"],
    }


def raw_artifact_paths() -> list[Path]:
    ensure_raw_artifacts()
    paths = sorted(path for path in email_raw_root().rglob("*") if path.is_file())
    return paths + [metadata_path(), ground_truth_path()]


def email_data_relative(path: Path) -> str:
    return str(path.resolve().relative_to(email_data_root().resolve())).replace("\\", "/")
