from __future__ import annotations

import csv
import hashlib
import json
import math
import random
import shutil
import textwrap
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

from app.data_paths import get_use_case_data_dir

USE_CASE_SLUG = "liquidity-forecast"
GENERATION_SEED = 5505
HISTORY_DAYS = 180
FORECAST_HORIZON_DAYS = 14
HISTORY_START_DATE = date(2026, 1, 1)
FILE_BASENAME = "synthetic_liquidity_cash_timeseries"

TIME_SERIES_HEADERS = [
    "series_id",
    "location_id",
    "location_name",
    "location_type",
    "region",
    "date",
    "day_of_week",
    "is_weekend",
    "holiday_flag",
    "holiday_name",
    "campaign_flag",
    "campaign_name",
    "cash_outflow",
    "cash_inflow",
    "net_cash_demand",
    "opening_cash",
    "closing_cash",
    "replenishment_amount",
    "stockout_event",
    "cash_capacity",
    "minimum_cash_threshold",
]

LOCATIONS: list[dict[str, Any]] = [
    {
        "series_id": "BR-CENTRAL-001",
        "location_id": "BR-001",
        "location_name": "Central City Branch",
        "location_type": "branch",
        "region": "Metro",
        "base_outflow": 116000,
        "base_inflow": 48500,
        "cash_capacity": 760000,
        "minimum_cash_threshold": 180000,
        "service_target": 0.98,
    },
    {
        "series_id": "BR-LAKESIDE-002",
        "location_id": "BR-002",
        "location_name": "Lakeside Branch",
        "location_type": "branch",
        "region": "North",
        "base_outflow": 87000,
        "base_inflow": 39200,
        "cash_capacity": 610000,
        "minimum_cash_threshold": 140000,
        "service_target": 0.97,
    },
    {
        "series_id": "BR-HARBOR-003",
        "location_id": "BR-003",
        "location_name": "Harbor Business Branch",
        "location_type": "branch",
        "region": "West",
        "base_outflow": 101000,
        "base_inflow": 53600,
        "cash_capacity": 690000,
        "minimum_cash_threshold": 155000,
        "service_target": 0.98,
    },
    {
        "series_id": "ATM-CENTRAL-101",
        "location_id": "ATM-101",
        "location_name": "Central Transit ATM",
        "location_type": "atm",
        "region": "Metro",
        "base_outflow": 35500,
        "base_inflow": 3900,
        "cash_capacity": 188000,
        "minimum_cash_threshold": 38000,
        "service_target": 0.96,
    },
    {
        "series_id": "ATM-MARKET-102",
        "location_id": "ATM-102",
        "location_name": "North Market ATM",
        "location_type": "atm",
        "region": "North",
        "base_outflow": 28800,
        "base_inflow": 3100,
        "cash_capacity": 156000,
        "minimum_cash_threshold": 32000,
        "service_target": 0.95,
    },
    {
        "series_id": "ATM-RESORT-103",
        "location_id": "ATM-103",
        "location_name": "West Resort ATM",
        "location_type": "atm",
        "region": "West",
        "base_outflow": 32600,
        "base_inflow": 2600,
        "cash_capacity": 172000,
        "minimum_cash_threshold": 35000,
        "service_target": 0.95,
    },
]

HOLIDAYS: list[dict[str, Any]] = [
    {
        "event_id": "HOL-2026-001",
        "event_type": "holiday",
        "name": "New Year Bank Holiday",
        "start_date": "2026-01-01",
        "end_date": "2026-01-01",
        "impact_multiplier": 0.55,
        "affected_location_type": "branch",
        "affected_region": "all",
    },
    {
        "event_id": "HOL-2026-002",
        "event_type": "holiday",
        "name": "Spring Service Holiday",
        "start_date": "2026-03-20",
        "end_date": "2026-03-20",
        "impact_multiplier": 0.68,
        "affected_location_type": "branch",
        "affected_region": "all",
    },
    {
        "event_id": "HOL-2026-003",
        "event_type": "holiday",
        "name": "Independence Day Synthetic Closure",
        "start_date": "2026-07-04",
        "end_date": "2026-07-04",
        "impact_multiplier": 0.62,
        "affected_location_type": "branch",
        "affected_region": "all",
    },
    {
        "event_id": "HOL-2026-004",
        "event_type": "holiday",
        "name": "Independence Day ATM Surge",
        "start_date": "2026-07-03",
        "end_date": "2026-07-05",
        "impact_multiplier": 1.28,
        "affected_location_type": "atm",
        "affected_region": "all",
    },
]

CAMPAIGNS: list[dict[str, Any]] = [
    {
        "event_id": "CMP-2026-001",
        "event_type": "campaign",
        "name": "Metro Payroll Bonus Campaign",
        "start_date": "2026-02-13",
        "end_date": "2026-02-18",
        "impact_multiplier": 1.17,
        "affected_location_type": "all",
        "affected_region": "Metro",
    },
    {
        "event_id": "CMP-2026-002",
        "event_type": "campaign",
        "name": "Spring Travel Cash Campaign",
        "start_date": "2026-05-22",
        "end_date": "2026-05-29",
        "impact_multiplier": 1.2,
        "affected_location_type": "atm",
        "affected_region": "all",
    },
    {
        "event_id": "CMP-2026-003",
        "event_type": "campaign",
        "name": "Summer Fee Waiver Campaign",
        "start_date": "2026-07-01",
        "end_date": "2026-07-10",
        "impact_multiplier": 1.12,
        "affected_location_type": "all",
        "affected_region": "all",
    },
]


def liquidity_data_root() -> Path:
    return get_use_case_data_dir(USE_CASE_SLUG)


def liquidity_raw_root() -> Path:
    return liquidity_data_root() / "raw"


def metadata_path() -> Path:
    return liquidity_data_root() / "metadata.json"


def ground_truth_path() -> Path:
    return liquidity_data_root() / "ground_truth.json"


def time_series_xlsx_path() -> Path:
    return liquidity_raw_root() / "timeseries" / f"{FILE_BASENAME}.xlsx"


def holiday_calendar_path() -> Path:
    return liquidity_raw_root() / "calendar" / "holiday_calendar.csv"


def campaign_calendar_path() -> Path:
    return liquidity_raw_root() / "calendar" / "campaign_calendar.csv"


def cash_policy_pdf_path() -> Path:
    return liquidity_raw_root() / "policies" / "cash_inventory_policy.pdf"


def _clean_raw() -> None:
    root = liquidity_raw_root()
    root.mkdir(parents=True, exist_ok=True)
    for path in root.iterdir():
        if path.name == ".gitkeep":
            continue
        if path.is_dir():
            shutil.rmtree(path)
        elif path.is_file():
            path.unlink()


def _date_range(start: date, days: int) -> list[date]:
    return [start + timedelta(days=offset) for offset in range(days)]


def _event_applies(event: dict[str, Any], location: dict[str, Any], current_date: date) -> bool:
    start = date.fromisoformat(event["start_date"])
    end = date.fromisoformat(event["end_date"])
    if not (start <= current_date <= end):
        return False
    event_location_type = event["affected_location_type"]
    event_region = event["affected_region"]
    return (
        event_location_type in {"all", location["location_type"]}
        and event_region in {"all", location["region"]}
    )


def _active_event(events: list[dict[str, Any]], location: dict[str, Any], current_date: date) -> dict[str, Any] | None:
    for event in events:
        if _event_applies(event, location, current_date):
            return event
    return None


def _day_factor(location_type: str, current_date: date) -> float:
    weekday = current_date.weekday()
    if location_type == "branch":
        return [1.02, 0.97, 1.0, 1.04, 1.16, 0.34, 0.18][weekday]
    return [0.93, 0.96, 1.0, 1.08, 1.23, 1.32, 1.18][weekday]


def _payday_factor(current_date: date) -> float:
    if current_date.day in {1, 2, 3, 15, 16}:
        return 1.18
    if current_date.day in {28, 29, 30, 31}:
        return 1.08
    return 1.0


def _season_factor(current_date: date) -> float:
    month_factor = {
        1: 0.92,
        2: 0.95,
        3: 1.0,
        4: 1.03,
        5: 1.08,
        6: 1.12,
        7: 1.16,
    }
    return month_factor.get(current_date.month, 1.0)


def _round_money(value: float) -> float:
    return round(max(0.0, value), 2)


def _sample_cash_record(
    location: dict[str, Any],
    current_date: date,
    opening_cash: float,
    rng: random.Random,
) -> dict[str, Any]:
    holiday = _active_event(HOLIDAYS, location, current_date)
    campaign = _active_event(CAMPAIGNS, location, current_date)
    trend_day = (current_date - HISTORY_START_DATE).days
    trend_factor = 1.0 + min(trend_day, HISTORY_DAYS + FORECAST_HORIZON_DAYS) * 0.00055
    multiplier = _day_factor(location["location_type"], current_date) * _payday_factor(current_date) * _season_factor(current_date) * trend_factor
    if holiday:
        multiplier *= float(holiday["impact_multiplier"])
    if campaign:
        multiplier *= float(campaign["impact_multiplier"])
    noise = rng.gauss(1.0, 0.055 if location["location_type"] == "branch" else 0.075)
    outflow = _round_money(location["base_outflow"] * multiplier * noise)
    inflow_noise = rng.gauss(1.0, 0.08)
    inflow = _round_money(location["base_inflow"] * (0.92 + 0.12 * math.sin(trend_day / 19)) * inflow_noise)
    if location["location_type"] == "atm":
        inflow = _round_money(inflow * (0.65 if current_date.weekday() >= 5 else 1.0))
    net_demand = _round_money(max(outflow - inflow, outflow * 0.36))
    projected_closing = opening_cash - net_demand
    replenishment = 0.0
    if projected_closing < location["minimum_cash_threshold"]:
        target_cash = location["cash_capacity"] * (0.78 if location["location_type"] == "branch" else 0.72)
        replenishment = _round_money(target_cash - projected_closing)
        projected_closing += replenishment
    stockout_event = 1 if projected_closing < location["minimum_cash_threshold"] * 0.7 else 0
    return {
        "series_id": location["series_id"],
        "location_id": location["location_id"],
        "location_name": location["location_name"],
        "location_type": location["location_type"],
        "region": location["region"],
        "date": current_date.isoformat(),
        "day_of_week": current_date.strftime("%A"),
        "is_weekend": 1 if current_date.weekday() >= 5 else 0,
        "holiday_flag": 1 if holiday else 0,
        "holiday_name": holiday["name"] if holiday else None,
        "campaign_flag": 1 if campaign else 0,
        "campaign_name": campaign["name"] if campaign else None,
        "cash_outflow": outflow,
        "cash_inflow": inflow,
        "net_cash_demand": net_demand,
        "opening_cash": _round_money(opening_cash),
        "closing_cash": _round_money(projected_closing),
        "replenishment_amount": replenishment,
        "stockout_event": stockout_event,
        "cash_capacity": float(location["cash_capacity"]),
        "minimum_cash_threshold": float(location["minimum_cash_threshold"]),
    }


def build_time_series(seed: int = GENERATION_SEED) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    rng = random.Random(seed)
    history_dates = _date_range(HISTORY_START_DATE, HISTORY_DAYS)
    holdout_start = HISTORY_START_DATE + timedelta(days=HISTORY_DAYS)
    holdout_dates = _date_range(holdout_start, FORECAST_HORIZON_DAYS)
    history: list[dict[str, Any]] = []
    holdout: list[dict[str, Any]] = []
    for location in LOCATIONS:
        opening_cash = location["cash_capacity"] * rng.uniform(0.58, 0.72)
        for current_date in history_dates:
            record = _sample_cash_record(location, current_date, opening_cash, rng)
            history.append(record)
            opening_cash = record["closing_cash"]
        for current_date in holdout_dates:
            record = _sample_cash_record(location, current_date, opening_cash, rng)
            holdout.append(record)
            opening_cash = record["closing_cash"]
    return history, holdout


def _write_time_series_workbook(history: list[dict[str, Any]], holdout: list[dict[str, Any]]) -> None:
    path = time_series_xlsx_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.properties.created = datetime(2026, 1, 1)
    workbook.properties.modified = datetime(2026, 1, 1)
    history_sheet = workbook.active
    history_sheet.title = "history"
    history_sheet.append(TIME_SERIES_HEADERS)
    for row in history:
        history_sheet.append([row.get(header) for header in TIME_SERIES_HEADERS])
    holdout_sheet = workbook.create_sheet("holdout_actuals")
    holdout_sheet.append(TIME_SERIES_HEADERS)
    for row in holdout:
        holdout_sheet.append([row.get(header) for header in TIME_SERIES_HEADERS])
    location_sheet = workbook.create_sheet("locations")
    location_headers = [
        "series_id",
        "location_id",
        "location_name",
        "location_type",
        "region",
        "cash_capacity",
        "minimum_cash_threshold",
        "service_target",
    ]
    location_sheet.append(location_headers)
    for location in LOCATIONS:
        location_sheet.append([location[header] for header in location_headers])
    for sheet in workbook.worksheets:
        for column in sheet.columns:
            letter = column[0].column_letter
            sheet.column_dimensions[letter].width = min(max(len(str(cell.value or "")) for cell in column) + 2, 32)
    workbook.save(path)


def _write_calendar_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    headers = [
        "event_id",
        "event_type",
        "name",
        "start_date",
        "end_date",
        "impact_multiplier",
        "affected_location_type",
        "affected_region",
    ]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def _write_cash_policy_pdf(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    content = [
        "Synthetic Cash Inventory Policy",
        "",
        "Purpose: define deterministic cash ordering controls for the Liquidity Forecast use case.",
        "Branch locations target a 98 percent service level, while ATM locations target a 95 to 96 percent service level.",
        "Forecast recommendations should prioritize stockout prevention, cash capacity limits, and operational delivery windows.",
        "Daily cash demand is reviewed with holiday, campaign, payday, weekend, and location-type adjustments.",
        "Replenishment should be recommended when projected closing cash is below the minimum threshold or when p90 demand creates material service risk.",
        "All values in this document are synthetic and are safe for local testing.",
    ]
    pdf = canvas.Canvas(str(path), pagesize=letter, invariant=1)
    pdf.setTitle("Synthetic Cash Inventory Policy")
    pdf.setFont("Helvetica-Bold", 15)
    pdf.drawString(54, 770, content[0])
    pdf.setFont("Helvetica", 10)
    y = 740
    for paragraph in content[2:]:
        for line in textwrap.wrap(paragraph, width=95):
            pdf.drawString(54, y, line)
            y -= 16
    pdf.save()


def _checksum(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _artifact_manifest(path: Path, artifact_type: str, group: str) -> dict[str, Any]:
    return {
        "file_name": path.name,
        "relative_path": str(path.relative_to(liquidity_data_root())).replace("\\", "/"),
        "artifact_type": artifact_type,
        "group": group,
        "checksum": _checksum(path),
    }


def write_artifacts() -> dict[str, str]:
    liquidity_data_root().mkdir(parents=True, exist_ok=True)
    _clean_raw()
    history, holdout = build_time_series()
    _write_time_series_workbook(history, holdout)
    _write_calendar_csv(holiday_calendar_path(), HOLIDAYS)
    _write_calendar_csv(campaign_calendar_path(), CAMPAIGNS)
    _write_cash_policy_pdf(cash_policy_pdf_path())
    artifacts = [
        _artifact_manifest(time_series_xlsx_path(), "xlsx", "timeseries"),
        _artifact_manifest(holiday_calendar_path(), "csv", "calendar"),
        _artifact_manifest(campaign_calendar_path(), "csv", "calendar"),
        _artifact_manifest(cash_policy_pdf_path(), "pdf", "policies"),
    ]
    holdout_start = HISTORY_START_DATE + timedelta(days=HISTORY_DAYS)
    metadata = {
        "dataset": FILE_BASENAME,
        "generation_seed": GENERATION_SEED,
        "history_start_date": HISTORY_START_DATE.isoformat(),
        "history_end_date": (holdout_start - timedelta(days=1)).isoformat(),
        "forecast_start_date": holdout_start.isoformat(),
        "forecast_end_date": (holdout_start + timedelta(days=FORECAST_HORIZON_DAYS - 1)).isoformat(),
        "history_days": HISTORY_DAYS,
        "forecast_horizon_days": FORECAST_HORIZON_DAYS,
        "location_count": len(LOCATIONS),
        "history_record_count": len(history),
        "holdout_record_count": len(holdout),
        "columns": TIME_SERIES_HEADERS,
        "locations": [
            {
                "series_id": location["series_id"],
                "location_id": location["location_id"],
                "location_name": location["location_name"],
                "location_type": location["location_type"],
                "region": location["region"],
                "cash_capacity": location["cash_capacity"],
                "minimum_cash_threshold": location["minimum_cash_threshold"],
                "service_target": location["service_target"],
            }
            for location in LOCATIONS
        ],
        "expected_topics": [
            "branch cash demand",
            "ATM replenishment",
            "holiday demand adjustment",
            "campaign demand adjustment",
            "stockout risk",
            "cash capacity control",
        ],
        "artifacts": artifacts,
        "description": "Synthetic branch and ATM liquidity time-series data with deterministic holdout actuals for forecast validation.",
    }
    ground_truth = {
        "generation_seed": GENERATION_SEED,
        "series_count": len(LOCATIONS),
        "history_days": HISTORY_DAYS,
        "forecast_horizon_days": FORECAST_HORIZON_DAYS,
        "holdout_actual_count": len(holdout),
        "expected_series_ids": [location["series_id"] for location in LOCATIONS],
        "actuals": [
            {
                "series_id": row["series_id"],
                "location_id": row["location_id"],
                "date": row["date"],
                "actual_net_cash_demand": row["net_cash_demand"],
                "holiday_flag": row["holiday_flag"],
                "campaign_flag": row["campaign_flag"],
            }
            for row in holdout
        ],
    }
    metadata_path().write_text(json.dumps(metadata, indent=2), encoding="utf-8")
    ground_truth_path().write_text(json.dumps(ground_truth, indent=2), encoding="utf-8")
    return {
        "raw_root": str(liquidity_raw_root().resolve()),
        "timeseries_xlsx": str(time_series_xlsx_path().resolve()),
        "holiday_calendar": str(holiday_calendar_path().resolve()),
        "campaign_calendar": str(campaign_calendar_path().resolve()),
        "cash_policy_pdf": str(cash_policy_pdf_path().resolve()),
        "metadata": str(metadata_path().resolve()),
        "ground_truth": str(ground_truth_path().resolve()),
    }
