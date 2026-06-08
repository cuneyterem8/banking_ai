import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.use_cases.liquidity_forecast.data_generation import (
    campaign_calendar_path,
    cash_policy_pdf_path,
    ground_truth_path,
    holiday_calendar_path,
    liquidity_data_root,
    liquidity_raw_root,
    metadata_path,
    time_series_xlsx_path,
    write_artifacts,
)
from app.use_cases.liquidity_forecast.schemas import (
    LiquidityCalendarEvent,
    LiquidityLocation,
    LiquidityTimeSeriesRecord,
)

USE_CASE_SLUG = "liquidity-forecast"
DATASET_KEY_CASH_TIMESERIES = "cash_timeseries"


def ensure_raw_artifacts() -> None:
    required = [
        time_series_xlsx_path(),
        holiday_calendar_path(),
        campaign_calendar_path(),
        cash_policy_pdf_path(),
        metadata_path(),
        ground_truth_path(),
    ]
    if not all(path.exists() for path in required):
        write_artifacts()
        return
    manifest = load_manifest(regenerate_if_missing=False)
    if not manifest.get("locations") or not manifest.get("artifacts"):
        write_artifacts()


def load_manifest(*, regenerate_if_missing: bool = True) -> dict[str, Any]:
    if regenerate_if_missing and not metadata_path().exists():
        write_artifacts()
    return json.loads(metadata_path().read_text(encoding="utf-8"))


def load_ground_truth() -> dict[str, Any]:
    ensure_raw_artifacts()
    return json.loads(ground_truth_path().read_text(encoding="utf-8"))


def _load_sheet_records(sheet_name: str) -> list[dict[str, Any]]:
    ensure_raw_artifacts()
    workbook = load_workbook(time_series_xlsx_path(), read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value) for value in rows[0]]
    records: list[dict[str, Any]] = []
    for values in rows[1:]:
        records.append({header: value for header, value in zip(headers, values)})
    workbook.close()
    return records


def load_history_records() -> list[LiquidityTimeSeriesRecord]:
    return [LiquidityTimeSeriesRecord(**record) for record in _load_sheet_records("history")]


def load_holdout_records() -> list[LiquidityTimeSeriesRecord]:
    return [LiquidityTimeSeriesRecord(**record) for record in _load_sheet_records("holdout_actuals")]


def load_locations() -> list[LiquidityLocation]:
    return [LiquidityLocation(**item) for item in load_manifest()["locations"]]


def _load_calendar(path: Path) -> list[LiquidityCalendarEvent]:
    ensure_raw_artifacts()
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        rows = []
        for row in reader:
            row["impact_multiplier"] = float(row["impact_multiplier"])
            rows.append(LiquidityCalendarEvent(**row))
        return rows


def load_calendar_events() -> list[LiquidityCalendarEvent]:
    return _load_calendar(holiday_calendar_path()) + _load_calendar(campaign_calendar_path())


def raw_artifact_paths() -> list[Path]:
    ensure_raw_artifacts()
    artifact_paths = sorted(path for path in liquidity_raw_root().rglob("*") if path.is_file())
    return artifact_paths + [metadata_path(), ground_truth_path()]


def manifest_preview(limit: int = 12) -> list[dict[str, Any]]:
    history = load_history_records()
    return [item.model_dump() for item in history[:limit]]


def location_preview() -> list[dict[str, Any]]:
    return [item.model_dump() for item in load_locations()]


def ground_truth_summary() -> dict[str, Any]:
    ground_truth = load_ground_truth()
    return {
        "series_count": ground_truth["series_count"],
        "history_days": ground_truth["history_days"],
        "forecast_horizon_days": ground_truth["forecast_horizon_days"],
        "holdout_actual_count": ground_truth["holdout_actual_count"],
        "expected_series_ids": ground_truth["expected_series_ids"],
    }
