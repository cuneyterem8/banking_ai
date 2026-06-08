from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.use_cases.market_intelligence.data_generation import (
    CALENDAR_EVENT_COUNT,
    COMPETITOR_RATE_COUNT,
    EVALUATION_CASE_COUNT,
    NEWS_COUNT,
    RATE_DAY_COUNT,
    calendar_path,
    competitors_path,
    evaluation_cases_path,
    ground_truth_path,
    market_data_root,
    market_raw_root,
    metadata_path,
    news_path,
    rates_path,
    snapshot_pdf_path,
    taxonomy_path,
    write_artifacts,
)
from app.use_cases.market_intelligence.schemas import (
    MarketResearchRequest,
    SyntheticCalendarEvent,
    SyntheticCompetitorRate,
    SyntheticMarketArticle,
    SyntheticRateRecord,
)

USE_CASE_SLUG = "market-intelligence"
DATASET_KEY_MARKET_INPUTS = "market_intelligence_inputs"


def ensure_raw_artifacts() -> None:
    required = [
        news_path(),
        rates_path(),
        competitors_path(),
        calendar_path(),
        snapshot_pdf_path(),
        taxonomy_path(),
        evaluation_cases_path(),
        metadata_path(),
        ground_truth_path(),
    ]
    if not all(path.exists() for path in required):
        write_artifacts()
        return
    manifest = load_manifest(regenerate_if_missing=False)
    if (
        manifest.get("news_count") != NEWS_COUNT
        or manifest.get("rate_record_count") != RATE_DAY_COUNT
        or manifest.get("competitor_rate_count") != COMPETITOR_RATE_COUNT
        or manifest.get("calendar_event_count") != CALENDAR_EVENT_COUNT
        or manifest.get("evaluation_case_count") != EVALUATION_CASE_COUNT
    ):
        write_artifacts()


def load_manifest(*, regenerate_if_missing: bool = True) -> dict[str, Any]:
    if regenerate_if_missing and not metadata_path().exists():
        write_artifacts()
    return json.loads(metadata_path().read_text(encoding="utf-8"))


def load_ground_truth(*, regenerate_if_missing: bool = True) -> dict[str, Any]:
    if regenerate_if_missing:
        ensure_raw_artifacts()
    return json.loads(ground_truth_path().read_text(encoding="utf-8"))


def load_news() -> list[SyntheticMarketArticle]:
    ensure_raw_artifacts()
    return [SyntheticMarketArticle(**item) for item in json.loads(news_path().read_text(encoding="utf-8"))]


def _load_csv_records(path: Path) -> list[dict[str, Any]]:
    with path.open("r", newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def load_rates() -> list[SyntheticRateRecord]:
    ensure_raw_artifacts()
    rows = []
    for item in _load_csv_records(rates_path()):
        rows.append(
            SyntheticRateRecord(
                date=item["date"],
                fed_funds_rate=float(item["fed_funds_rate"]),
                treasury_10y=float(item["treasury_10y"]),
                mortgage_30y=float(item["mortgage_30y"]),
                deposit_beta_index=float(item["deposit_beta_index"]),
                usd_index=float(item["usd_index"]),
                inflation_expectation=float(item["inflation_expectation"]),
            )
        )
    return rows


def _load_xlsx_records(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if len(rows) < 2:
        return []
    headers = [str(item) for item in rows[0]]
    return [
        {header: value for header, value in zip(headers, row)}
        for row in rows[1:]
        if any(value is not None for value in row)
    ]


def load_competitor_rates() -> list[SyntheticCompetitorRate]:
    ensure_raw_artifacts()
    return [SyntheticCompetitorRate(**item) for item in _load_xlsx_records(competitors_path())]


def load_calendar_events() -> list[SyntheticCalendarEvent]:
    ensure_raw_artifacts()
    rows: list[SyntheticCalendarEvent] = []
    for item in _load_csv_records(calendar_path()):
        rows.append(
            SyntheticCalendarEvent(
                event_id=item["event_id"],
                event_date=item["event_date"],
                event_type=item["event_type"],
                title=item["title"],
                expected_impact=item["expected_impact"],
                affected_areas=[value for value in item["affected_areas"].split("|") if value],
            )
        )
    return rows


def load_taxonomy() -> dict[str, Any]:
    ensure_raw_artifacts()
    return json.loads(taxonomy_path().read_text(encoding="utf-8"))


def load_evaluation_cases() -> list[MarketResearchRequest]:
    ensure_raw_artifacts()
    return [MarketResearchRequest(**item) for item in json.loads(evaluation_cases_path().read_text(encoding="utf-8"))]


def manifest_preview(limit: int = 16) -> list[dict[str, Any]]:
    preview: list[dict[str, Any]] = []
    preview.extend(
        {
            "record_type": "news",
            "id": item.article_id,
            "topic": item.topic,
            "impact_area": item.impact_area,
            "sentiment": item.sentiment,
        }
        for item in load_news()[:5]
    )
    preview.extend(
        {
            "record_type": "rate",
            "id": item.date,
            "fed_funds_rate": item.fed_funds_rate,
            "treasury_10y": item.treasury_10y,
            "mortgage_30y": item.mortgage_30y,
        }
        for item in load_rates()[-5:]
    )
    preview.extend(
        {
            "record_type": "competitor",
            "id": item.competitor_id,
            "product_line": item.product_line,
            "rate": item.rate,
        }
        for item in load_competitor_rates()[:4]
    )
    preview.extend(
        {
            "record_type": "evaluation_case",
            "id": f"case_{index + 1}",
            "objective": item.objective,
            "depth": item.depth,
        }
        for index, item in enumerate(load_evaluation_cases()[:2])
    )
    return preview[:limit]


def ground_truth_summary() -> dict[str, Any]:
    ground_truth = load_ground_truth()
    return {
        "news_count": ground_truth["news_count"],
        "rate_record_count": ground_truth["rate_record_count"],
        "competitor_rate_count": ground_truth["competitor_rate_count"],
        "calendar_event_count": ground_truth["calendar_event_count"],
        "evaluation_case_count": ground_truth["evaluation_case_count"],
        "expected_topics": ground_truth["expected_topics"],
        "expected_impact_areas": ground_truth["expected_impact_areas"],
    }


def raw_artifact_paths() -> list[Path]:
    ensure_raw_artifacts()
    paths = sorted(path for path in market_raw_root().rglob("*") if path.is_file())
    return paths + [metadata_path(), ground_truth_path()]


def market_data_relative(path: Path) -> str:
    return str(path.resolve().relative_to(market_data_root().resolve())).replace("\\", "/")
